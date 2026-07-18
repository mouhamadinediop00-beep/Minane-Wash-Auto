# -*- coding: utf-8 -*-
"""
LAVAGE MECKHE - Module exports
MODULE 5  : factures, tickets, reçus, devis (PDF + QR code)
MODULE 12 : rapports Excel / PDF / CSV
MODULE 13 : classeur Excel automatique multi-feuilles
"""

import os
import csv
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                Spacer, Image as RLImage)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

import qrcode

from .database import get_conn, get_parametre, BASE_DIR
from .services import (stock_reel, journal_de_caisse, calculer_kpi,
                       historique_clients, historique_vehicules,
                       prochain_numero, FCFA)

DOSSIER_EXPORTS = os.path.join(BASE_DIR, "exports")
DOSSIER_FACTURES = os.path.join(BASE_DIR, "factures")
DOSSIER_ASSETS = os.path.join(BASE_DIR, "assets")
os.makedirs(DOSSIER_EXPORTS, exist_ok=True)
os.makedirs(DOSSIER_FACTURES, exist_ok=True)
os.makedirs(DOSSIER_ASSETS, exist_ok=True)


def _dossier_mois(base: str, jour=None) -> str:
    """Retourne (et crée) un sous-dossier AAAA-MM pour ranger les fichiers par mois."""
    d = jour or date.today()
    chemin = os.path.join(base, f"{d:%Y-%m}")
    os.makedirs(chemin, exist_ok=True)
    return chemin


def trouver_logo() -> str:
    """
    Renvoie le chemin du logo à utiliser sur les factures, ou "" si aucun.
    Priorité : chemin défini dans Paramètres, sinon détection automatique
    d'un fichier « logo.* » déposé dans le dossier assets.
    """
    perso = get_parametre("entreprise_logo", "")
    if perso and os.path.exists(perso):
        return perso
    for nom in ("logo.png", "logo.jpg", "logo.jpeg", "logo.svg",
                "logo.PNG", "logo.JPG", "logo.SVG"):
        chemin = os.path.join(DOSSIER_ASSETS, nom)
        if os.path.exists(chemin):
            return chemin
    return ""


def _logo_flowable(chemin: str, largeur_mm: float = 34):
    """
    Prépare le logo pour reportlab (PNG/JPG en Image, SVG converti en dessin),
    en conservant les proportions. Renvoie un flowable ou None.
    """
    if not chemin or not os.path.exists(chemin):
        return None
    ext = os.path.splitext(chemin)[1].lower()
    largeur = largeur_mm * mm
    try:
        if ext == ".svg":
            from svglib.svglib import svg2rlg
            dessin = svg2rlg(chemin)
            if dessin is None:
                return None
            facteur = largeur / dessin.width if dessin.width else 1
            dessin.width *= facteur
            dessin.height *= facteur
            dessin.scale(facteur, facteur)
            return dessin
        else:
            from PIL import Image as PILImage
            with PILImage.open(chemin) as im:
                ratio = im.height / im.width if im.width else 1
            return RLImage(chemin, width=largeur, height=largeur * ratio)
    except Exception:
        return None

# --- Styles Excel -----------------------------------------------------
BLEU = "1B4F72"
FOND_ENTETE = PatternFill("solid", fgColor=BLEU)
POLICE_ENTETE = Font(color="FFFFFF", bold=True, size=11)
BORDURE = Border(*[Side(style="thin", color="B0B0B0")] * 4)


def _feuille(wb, titre, entetes, lignes, largeurs=None):
    ws = wb.create_sheet(titre)
    ws.append(entetes)
    for c in ws[1]:
        c.fill, c.font, c.border = FOND_ENTETE, POLICE_ENTETE, BORDURE
        c.alignment = Alignment(horizontal="center")
    for ligne in lignes:
        ws.append(list(ligne))
    for i, e in enumerate(entetes, 1):
        ws.column_dimensions[get_column_letter(i)].width = (largeurs or {}).get(i, max(len(str(e)) + 4, 14))
    ws.freeze_panes = "A2"
    return ws


# ======================================================================
# MODULE 13 : CLASSEUR EXCEL COMPLET (feuilles liées)
# ======================================================================
def exporter_classeur_complet(chemin=None) -> str:
    """Génère le classeur Excel avec toutes les feuilles + tableau de bord KPI lié."""
    chemin = chemin or os.path.join(
        DOSSIER_EXPORTS, f"GESTION_LAVAGE_{date.today():%Y-%m-%d}.xlsx")
    conn = get_conn()
    wb = Workbook()
    wb.remove(wb.active)

    # --- Ventes ---
    ventes = conn.execute("""
        SELECT v.numero, v.date, v.heure, u.nom, c.nom, ve.plaque,
               v.montant_brut, v.remise, v.montant_net, v.mode_paiement, v.statut
        FROM ventes v LEFT JOIN utilisateurs u ON u.id=v.caissier_id
        LEFT JOIN clients c ON c.id=v.client_id
        LEFT JOIN vehicules ve ON ve.id=v.vehicule_id ORDER BY v.date, v.heure
    """).fetchall()
    _feuille(wb, "Ventes",
             ["Numéro", "Date", "Heure", "Caissier", "Client", "Plaque",
              "Montant brut", "Remise", "Montant net", "Mode paiement", "Statut"],
             ventes)

    # --- Clients ---
    _feuille(wb, "Clients",
             ["Nom", "Téléphone", "Adresse", "Profession", "Entreprise", "Type",
              "Carte fidélité", "Visites", "Total dépensé", "Dernière visite"],
             [(c["nom"], c["telephone"], c["adresse"], c["profession"], c["entreprise"],
               c["type_client"], "Oui" if c["carte_fidelite"] else "Non",
               c["visites"], c["total"], c["derniere"]) for c in historique_clients()])

    # --- Prestations ---
    _feuille(wb, "Prestations",
             ["Code", "Nom", "Type véhicule", "Type lavage", "Prix", "Durée (min)", "TVA %"],
             conn.execute("SELECT code, nom, type_vehicule, type_lavage, prix, duree_min, tva "
                          "FROM prestations WHERE actif=1 ORDER BY code").fetchall())

    # --- Stock ---
    _feuille(wb, "Stock",
             ["Code", "Produit", "Catégorie", "Unité", "Stock initial", "Entrées",
              "Sorties", "Stock réel", "Stock min", "Prix achat", "Valeur stock",
              "Péremption", "Fournisseur", "Critique"],
             [(p["code"], p["nom"], p["categorie"], p["unite"], p["stock_initial"],
               p["entrees"], p["sorties"], p["stock_reel"], p["stock_min"],
               p["prix_achat"], p["valeur_stock"], p["date_peremption"],
               p["fournisseur"], "OUI" if p["critique"] else "") for p in stock_reel()])

    # --- Achats ---
    _feuille(wb, "Achats",
             ["Numéro", "Date", "Fournisseur", "Statut", "Total", "Payé"],
             conn.execute("SELECT a.numero, a.date, f.nom, a.statut, a.total, a.montant_paye "
                          "FROM achats a LEFT JOIN fournisseurs f ON f.id=a.fournisseur_id "
                          "ORDER BY a.date").fetchall())

    # --- Dépenses ---
    _feuille(wb, "Dépenses",
             ["Date", "Catégorie", "Libellé", "Montant", "Mode paiement"],
             conn.execute("SELECT date, categorie, libelle, montant, mode_paiement "
                          "FROM depenses ORDER BY date").fetchall())

    # --- Employés ---
    _feuille(wb, "Employés",
             ["Nom", "Téléphone", "Fonction", "Salaire", "Prime", "Horaires", "Actif"],
             conn.execute("SELECT nom, telephone, fonction, salaire, prime, horaires, "
                          "CASE actif WHEN 1 THEN 'Oui' ELSE 'Non' END FROM employes").fetchall())

    # --- Fournisseurs ---
    _feuille(wb, "Fournisseurs", ["Nom", "Téléphone", "Adresse"],
             conn.execute("SELECT nom, telephone, adresse FROM fournisseurs").fetchall())

    # --- Factures (documents émis) ---
    _feuille(wb, "Factures", ["Numéro", "Type", "Date", "Vente liée", "Fichier"],
             conn.execute("SELECT d.numero, d.type, d.date, v.numero, d.fichier "
                          "FROM documents d LEFT JOIN ventes v ON v.id=d.vente_id "
                          "ORDER BY d.date").fetchall())

    # --- Historique véhicules ---
    _feuille(wb, "Historique",
             ["Plaque", "Marque", "Modèle", "Couleur", "Catégorie", "Client",
              "Nb lavages", "Total", "Dernier lavage"],
             [(v["plaque"], v["marque"], v["modele"], v["couleur"], v["categorie"],
               v["client"], v["nb_lavages"], v["total"], v["dernier"])
              for v in historique_vehicules()])

    # --- Tableau de bord KPI : formules liées à la feuille Ventes ---
    ws = wb.create_sheet("Tableau de bord KPI", 0)
    ws["A1"] = get_parametre("entreprise_nom")
    ws["A1"].font = Font(bold=True, size=16, color=BLEU)
    ws["A2"] = f"Tableau de bord généré le {date.today():%d/%m/%Y}"
    nb = len(ventes) + 1
    formules = [
        ("CA total (FCFA)",        f"=SUMIF(Ventes!K2:K{nb},\"Payée\",Ventes!I2:I{nb})"),
        ("CA du jour (FCFA)",      f"=SUMIFS(Ventes!I2:I{nb},Ventes!B2:B{nb},TEXT(TODAY(),\"yyyy-mm-dd\"),Ventes!K2:K{nb},\"Payée\")"),
        ("Nombre de ventes",       f"=COUNTIF(Ventes!K2:K{nb},\"Payée\")"),
        ("Panier moyen (FCFA)",    "=IFERROR(B5/B7,0)"),
        ("Total dépenses (FCFA)",  "=SUM(Dépenses!D:D)"),
        ("Valeur du stock (FCFA)", "=SUM(Stock!K:K)"),
        ("Produits critiques",     "=COUNTIF(Stock!N:N,\"OUI\")"),
        ("Bénéfice estimé (FCFA)", "=B5-B9"),
    ]
    for i, (libelle, formule) in enumerate(formules, start=5):
        ws[f"A{i}"], ws[f"B{i}"] = libelle, formule
        ws[f"A{i}"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    conn.close()
    wb.save(chemin)
    return chemin


# ======================================================================
# MODULE 12 : RAPPORTS (Excel / PDF / CSV)
# ======================================================================
def exporter_rapport_kpi(debut: str, fin: str, formats=("xlsx", "pdf", "csv")) -> list:
    """Rapport de période (journalier/hebdo/mensuel/annuel selon les dates)."""
    kpi = calculer_kpi(debut, fin)
    from datetime import datetime as _dt
    _fin = _dt.fromisoformat(fin).date()
    base = os.path.join(_dossier_mois(DOSSIER_EXPORTS, _fin), f"RAPPORT_{debut}_au_{fin}")
    fichiers = []

    lignes = []
    for section in ("commercial", "exploitation", "stock", "finance", "marketing"):
        for k, v in kpi[section].items():
            lignes.append((section.capitalize(), k, v))

    if "xlsx" in formats:
        wb = Workbook()
        wb.remove(wb.active)
        _feuille(wb, "KPI", ["Section", "Indicateur", "Valeur"], lignes)
        _feuille(wb, "CA par prestation", ["Prestation", "CA", "Quantité"],
                 [(r["nom"], r["ca"], r["nb"]) for r in kpi["ca_par_prestation"]])
        _feuille(wb, "Revente produits", ["Produit", "Quantité vendue", "CA", "Marge (vente-achat)"],
                 [(r["nom"], r["nb"], r["ca"], r["marge"]) for r in kpi["ca_par_produit"]])
        _feuille(wb, "CA par employé", ["Employé", "CA", "Nb ventes"],
                 [(r["nom"], r["ca"], r["nb"]) for r in kpi["ca_par_employe"]])
        _feuille(wb, "Top 20 clients", ["Client", "Téléphone", "Visites", "Total"],
                 [(r["nom"], r["telephone"], r["visites"], r["total"]) for r in kpi["top_clients"]])
        f = base + ".xlsx"; wb.save(f); fichiers.append(f)

    if "csv" in formats:
        f = base + ".csv"
        with open(f, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh, delimiter=";")
            w.writerow(["Section", "Indicateur", "Valeur"])
            w.writerows(lignes)
        fichiers.append(f)

    if "pdf" in formats:
        f = base + ".pdf"
        doc = SimpleDocTemplate(f, pagesize=A4, topMargin=15 * mm)
        st = getSampleStyleSheet()
        titre = ParagraphStyle("t", parent=st["Title"], textColor=colors.HexColor("#1B4F72"))
        elems = [Paragraph(get_parametre("entreprise_nom"), titre),
                 Paragraph(f"Rapport d'activité du {debut} au {fin}", st["Heading2"]),
                 Spacer(1, 6 * mm)]
        data = [["Section", "Indicateur", "Valeur"]] + [
            [s, k, FCFA(v) if isinstance(v, (int, float)) and ("CA" in k or "FCFA" in k or
             "Marge" in k or "Bénéfice" in k or "Valeur" in k or "Coût" in k or
             "Charges" in k or "Panier" in k or "Seuil" in k or "Consommation" in k or
             "affaires" in k) else str(v)]
            for s, k, v in lignes]
        t = Table(data, colWidths=[35 * mm, 90 * mm, 45 * mm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F72")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2F8")]),
        ]))
        elems.append(t)
        doc.build(elems)
        fichiers.append(f)

    return fichiers


def exporter_journal_caisse_pdf(jour: str = None) -> str:
    """Journal de caisse de fin de journée (MODULE 4)."""
    j = journal_de_caisse(jour)
    from datetime import datetime as _dt
    _jour = _dt.fromisoformat(j['date']).date()
    f = os.path.join(_dossier_mois(DOSSIER_EXPORTS, _jour), f"JOURNAL_CAISSE_{j['date']}.pdf")
    doc = SimpleDocTemplate(f, pagesize=A4, topMargin=15 * mm)
    st = getSampleStyleSheet()
    elems = [Paragraph(get_parametre("entreprise_nom"), st["Title"]),
             Paragraph(f"Journal de caisse du {j['date']}", st["Heading2"]),
             Spacer(1, 4 * mm)]

    data = [["N°", "Heure", "Caissier", "Client", "Plaque", "Montant", "Paiement", "Statut"]]
    for v in j["ventes"]:
        data.append([v["numero"], v["heure"][:5], v["caissier"] or "-", v["client"] or "Passage",
                     v["plaque"] or "-", FCFA(v["montant_net"]), v["mode_paiement"], v["statut"]])
    t = Table(data, colWidths=[32 * mm, 13 * mm, 22 * mm, 30 * mm, 22 * mm, 24 * mm, 24 * mm, 18 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F72")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
    ]))
    elems += [t, Spacer(1, 6 * mm), Paragraph("Totaux par mode de paiement", st["Heading3"])]

    tot = [["Mode", "Montant"]] + [[m, FCFA(v)] for m, v in j["totaux"].items()]
    tot += [["TOTAL ENCAISSEMENTS", FCFA(j["ca"])],
            ["TOTAL DÉPENSES", FCFA(j["depenses"])],
            ["SOLDE DE CAISSE", FCFA(j["solde"])]]
    t2 = Table(tot, colWidths=[70 * mm, 50 * mm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F72")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTNAME", (0, -3), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D5F5E3")),
    ]))
    elems.append(t2)
    doc.build(elems)
    return f


# ======================================================================
# MODULE 5 : FACTURES / TICKETS / RECUS / DEVIS
# ======================================================================
def _qr_code(texte: str) -> str:
    chemin = os.path.join(DOSSIER_FACTURES, "_qr_tmp.png")
    qrcode.make(texte).save(chemin)
    return chemin


def generer_document(vente_id: int, type_doc: str = "Ticket") -> str:
    """Génère un PDF (Facture, Facture simplifiée, Ticket, Reçu ou Devis) pour une vente."""
    conn = get_conn()
    v = conn.execute(
        "SELECT v.*, c.nom client, c.telephone tel_client, c.adresse adr_client, "
        "ve.plaque, u.nom caissier FROM ventes v "
        "LEFT JOIN clients c ON c.id=v.client_id "
        "LEFT JOIN vehicules ve ON ve.id=v.vehicule_id "
        "LEFT JOIN utilisateurs u ON u.id=v.caissier_id WHERE v.id=?", (vente_id,)
    ).fetchone()
    if not v:
        conn.close()
        raise ValueError("Vente introuvable.")
    lignes = conn.execute(
        "SELECT p.nom, l.prix, l.quantite FROM vente_lignes l "
        "JOIN prestations p ON p.id=l.prestation_id WHERE l.vente_id=?", (vente_id,)
    ).fetchall()
    lignes_produits = conn.execute(
        "SELECT p.nom, vp.prix, vp.quantite FROM vente_produits vp "
        "JOIN produits p ON p.id=vp.produit_id WHERE vp.vente_id=?", (vente_id,)
    ).fetchall()

    prefixes = {"Facture": "FAC-", "Facture simplifiée": "FS-", "Ticket": "TK-",
                "Reçu": "RC-", "Devis": "DV-"}
    numero = prochain_numero(f"{prefixes[type_doc]}{date.today():%Y}-", "documents")
    fichier = os.path.join(_dossier_mois(DOSSIER_FACTURES), f"{numero}.pdf")

    ent = {k: get_parametre(f"entreprise_{k}")
           for k in ("nom", "ninea", "rccm", "adresse", "telephone", "telephone2", "email")}

    doc = SimpleDocTemplate(fichier, pagesize=A4, topMargin=15 * mm, bottomMargin=15 * mm)
    st = getSampleStyleSheet()
    bleu = colors.HexColor("#1B4F72")
    elems = []

    # En-tête entreprise : logo centré (détection auto dans assets/ ou chemin Paramètres)
    logo = _logo_flowable(trouver_logo(), largeur_mm=32)
    if logo is not None:
        try:
            logo.hAlign = "CENTER"
        except Exception:
            pass
        elems.append(logo)
        elems.append(Spacer(1, 3 * mm))
    style_nom = ParagraphStyle("e", parent=st["Title"], textColor=bleu, alignment=TA_CENTER)
    style_coord = ParagraphStyle("c", parent=st["Normal"], alignment=TA_CENTER, fontSize=9)
    elems.append(Paragraph(ent["nom"], style_nom))
    tels = ent["telephone"] + (f" / {ent['telephone2']}" if ent.get("telephone2") else "")
    elems.append(Paragraph(
        f"NINEA : {ent['ninea']} — RCCM : {ent['rccm']}<br/>"
        f"{ent['adresse']}<br/>"
        f"Tél : {tels} — {ent['email']}", style_coord))
    elems.append(Spacer(1, 5 * mm))
    elems.append(Paragraph(f"<b>{type_doc.upper()} N° {numero}</b>", st["Heading2"]))
    elems.append(Paragraph(
        f"Date : {v['date']} à {v['heure'][:5]} — Caissier : {v['caissier'] or '-'}<br/>"
        f"Client : {v['client'] or 'Client de passage'}"
        + (f" — Tél : {v['tel_client']}" if v["tel_client"] else "")
        + (f"<br/>Véhicule : {v['plaque']}" if v["plaque"] else ""), st["Normal"]))
    elems.append(Spacer(1, 4 * mm))

    data = [["Désignation", "Prix unitaire", "Qté", "Montant"]]
    for l in lignes:
        data.append([l["nom"], FCFA(l["prix"]), str(l["quantite"]), FCFA(l["prix"] * l["quantite"])])
    for l in lignes_produits:
        q = l["quantite"]
        q_txt = str(int(q)) if float(q).is_integer() else f"{q:g}"
        data.append([f"{l['nom']} (produit)", FCFA(l["prix"]), q_txt,
                     FCFA(l["prix"] * q)])
    nb_lignes_total = len(lignes) + len(lignes_produits)
    data += [["", "", "Sous-total", FCFA(v["montant_brut"])],
             ["", "", "Remise", FCFA(v["remise"])],
             ["", "", "NET À PAYER", FCFA(v["montant_net"])],
             ["", "", "Payé / Rendu",
              f"{FCFA(v['montant_paye'])} / {FCFA(v['monnaie_rendue'])}"],
             ["", "", "Mode", v["mode_paiement"]]]
    t = Table(data, colWidths=[80 * mm, 35 * mm, 25 * mm, 40 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), bleu),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, nb_lignes_total), 0.4, colors.grey),
        ("FONTNAME", (2, -3), (-1, -3), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 8 * mm))

    # QR Code : contient les références du document
    qr = _qr_code(f"{ent['nom']} | {type_doc} {numero} | {v['date']} | {FCFA(v['montant_net'])}")
    elems.append(RLImage(qr, width=25 * mm, height=25 * mm))
    elems.append(Paragraph("Merci de votre visite !", st["Italic"]))
    doc.build(elems)

    conn.execute(
        "INSERT INTO documents (numero, type, vente_id, fichier) VALUES (?,?,?,?)",
        (numero, type_doc, vente_id, fichier),
    )
    conn.commit()
    conn.close()
    return fichier


# ======================================================================
# JOURNAL D'AUDIT : export Excel
# ======================================================================
def exporter_audit(recherche="", debut=None, fin=None) -> str:
    """Exporte le journal d'audit filtré vers un fichier Excel."""
    from .audit import lister
    entrees = lister(limite=10000, recherche=recherche, debut=debut, fin=fin)
    wb = Workbook()
    wb.remove(wb.active)
    _feuille(wb, "Journal d'audit",
             ["Date", "Heure", "Utilisateur", "Action", "Détails"],
             [(e["date"], e["heure"], e["utilisateur"], e["action"], e["details"] or "")
              for e in entrees],
             largeurs={1: 12, 2: 12, 3: 20, 4: 28, 5: 50})
    chemin = os.path.join(_dossier_mois(DOSSIER_EXPORTS),
                          f"AUDIT_{date.today():%Y-%m-%d}.xlsx")
    wb.save(chemin)
    return chemin
